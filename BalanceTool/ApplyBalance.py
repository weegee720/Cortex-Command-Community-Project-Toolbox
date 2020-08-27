import csv
import sys
import os

from openpyxl import load_workbook
from openpyxl import Workbook

class CustomExcelDialect(csv.Dialect): 
    delimiter = ';' 
    doublequote = True
    escapechar = None
    lineterminator = '\r\n' 
    quotechar = '"' 
    quoting = csv.QUOTE_NONE
    skipinitialspace = True 

def LoadParametersFromExcel(input_file_name):
	result = dict()

	wb = load_workbook(input_file_name)
	sheet = wb["Parameters"]

	row = 2

	while sheet.cell(row = row, column = 1).value != None:
		preset_name = sheet.cell(row = row, column = 1).value
		parameter = sheet.cell(row = row, column = 2).value
		parameter_value = sheet.cell(row = row, column = 3).value
		
		if (parameter != None and parameter != "" and parameter_value != None and parameter_value != ""):
			full_parameter_name = preset_name + ":" + parameter
		
			result[full_parameter_name] = dict()
			result[full_parameter_name]["value"] = parameter_value
			result[full_parameter_name]["applied"] = False
			
			#print (preset_name, parameter, parameter_value)
		
		row = row + 1

	wb.close()

	return result

def LoadParametersFromCsv(input_file_name):
	result = dict()

	file = open(input_file_name, "r")
	csv_reader = csv.DictReader(file , dialect = CustomExcelDialect)

	for rec in csv_reader:
		#Entity;Parameter;Value;
		
		preset_name = rec["Entity"]
		parameter = rec["Parameter"]
		parameter_value = rec["Value"]
		
		if (parameter != None and parameter != "" and parameter_value != None and parameter_value != ""):
			full_parameter_name = preset_name + ":" + parameter
		
			result[full_parameter_name] = dict()
			result[full_parameter_name]["value"] = parameter_value
			result[full_parameter_name]["applied"] = False

			print (preset_name, parameter, parameter_value)

	file.close()

	return result

def ParseFile(parameters, data_folder, module_name, ini_file, indent):
	#print(ini_file)
	
	file_path = os.path.join(data_folder, ini_file)
	
	input = open(file_path, 'r')
	lines = input.readlines()
	
	curobject = dict()
	
	next_comment_mode = False
	comment_mode = False

	file_changed = False
	processed_lines = []
	current_preset = {}
	current_preset["preset"] = ""
	current_preset["indent"] = 0
	
	presets_stack = []
	
	for l in lines:
		#Discard comments
		cmnts = l.split("//")
		ln = cmnts[0]
		
		cmnts = ln.split("/*")
		ln = cmnts[0]

		if len(cmnts) > 1:
			next_comment_mode = True

		cmnts = ln.split("*/")
		if len(cmnts) > 1:
			ln = cmnts[1]
			next_comment_mode = False
			comment_mode = False

		if not comment_mode: 
			indent_level = 0
			
			while len(ln) > indent_level and ln[indent_level] == '\t':
				indent_level = indent_level + 1

			if indent_level < current_preset["indent"] and len(presets_stack) > 0:
				current_preset = presets_stack.pop()
		
			values = ln.split("=");
			
			if len(values) == 2:
				lvalue = values[0].strip()
				rvalue = values[1].strip()
				
				full_param = module_name + "/" + current_preset["preset"] + ":" + lvalue
				
				#print(full_param)
				
				if full_param in parameters:
					file_changed = True
					parameters[full_param]["applied"] = True
					
					l = ('\t' * indent_level) + lvalue + " = " + str(parameters[full_param]["value"]) + "\n"
					
					print ("+ " + full_param + " @ " + ini_file)
				
				if lvalue == "PresetName" and rvalue != "CopyOf":
					if current_preset["preset"] != "" and current_preset["indent"] < indent_level:
						presets_stack.append(current_preset)
					
					current_preset = {}
					current_preset["preset"] = rvalue
					current_preset["indent"] = indent_level
				
				if lvalue == "IncludeFile":
					ParseFile(parameters, data_folder, module_name, rvalue, True)
	
		comment_mode = next_comment_mode
		
		processed_lines.extend(l)

	input.close()

	if file_changed:
		print ("=> " + ini_file)
	
		output = open(file_path, "w", encoding='utf8')
		for l in processed_lines:
			output.write(l)
		output.close()

csv.register_dialect("custom_dialect", CustomExcelDialect)

if len(sys.argv) < 4:
	print ("")
	print ("")
	print ("!!! ALWAYS COMMIT YOUR INIs BEFORE USING THIS TOOL SO YOU CAN ROLL BACK IF IT GLITCHES !!!")
	print ("")
	print ("")
	print ("USAGE: ApplyBalance.py <data_directory> <module_name> <balance_csv_or_xlsx_file>")
	print ("")
	print ("EXAMPLE: ApplyBalance.py ..\..\Cortex-Command-Community-Project-Data Base.rte example_balance.xlsx")
	print ("")
	print ("")

	exit()

data_folder = sys.argv[1]
module_name = sys.argv[2]
balance_file = sys.argv[3]

parameters = dict()

if balance_file.endswith(".csv"):
	parameters = LoadParametersFromCsv(balance_file)
	
if balance_file.endswith(".xlsx"):
	parameters = LoadParametersFromExcel(balance_file)

ParseFile(parameters, data_folder, module_name, os.path.join(module_name, "Index.ini"), False)

all_applied = True

for key in parameters:
	if parameters[key]["applied"] == False:
		if all_applied:
			print ("\n")
			print ("NOT APPLIED:")
	
		all_applied = False
		print ("- "  + key)

if all_applied:
	print ("ALL CHANGES APPLIED")
